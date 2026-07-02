from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from whatsapp_core import normalize_phone

ALLOWED_EXTENSIONS = {".xlsx", ".xlsm"}


class ExcelImportError(Exception):
    def __init__(self, message: str, details: list[str] | None = None):
        super().__init__(message)
        self.message = message
        self.details = details or []


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_header(value: str) -> bool:
    lowered = value.lower()
    return lowered in {"phone", "number", "mobile", "contact", "phone number", "whatsapp"}


def parse_excel_phone_list(file_bytes: bytes, filename: str) -> list[str]:
    extension = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension not in ALLOWED_EXTENSIONS:
        raise ExcelImportError(
            "Only Excel files (.xlsx) with a single column of phone numbers are accepted.",
            [
                f"Received: {filename or 'unknown file'}",
                "Please export your list as .xlsx with one column only.",
            ],
        )

    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelImportError(
            "Could not read the Excel file. Make sure it is a valid .xlsx file.",
            [str(exc)],
        ) from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ExcelImportError("The Excel file is empty.")

    non_empty_columns: set[int] = set()
    parsed_rows: list[list[str]] = []

    for row in rows:
        values = [_cell_value(cell) for cell in row]
        while values and not values[-1]:
            values.pop()
        if not any(values):
            continue

        parsed_rows.append(values)
        for col_index, value in enumerate(values):
            if value:
                non_empty_columns.add(col_index)

    if not parsed_rows:
        raise ExcelImportError("The Excel file has no phone numbers.")

    if len(non_empty_columns) > 1:
        raise ExcelImportError(
            "The Excel file must contain only one column of phone numbers.",
            [
                f"Found data in {len(non_empty_columns)} columns.",
                "Remove extra columns and keep only phone numbers in column A.",
            ],
        )

    numbers: list[str] = []
    errors: list[str] = []
    start_index = 0

    first_value = parsed_rows[0][0] if parsed_rows[0] else ""
    if _looks_like_header(first_value):
        start_index = 1

    for row_number, row in enumerate(parsed_rows[start_index:], start=start_index + 1):
        if len(row) > 1 and any(row[1:]):
            raise ExcelImportError(
                "The Excel file must contain only one column of phone numbers.",
                [f"Extra data found on row {row_number}."],
            )

        raw = row[0] if row else ""
        if not raw:
            continue

        try:
            numbers.append(normalize_phone(raw))
        except ValueError as exc:
            errors.append(f"Row {row_number}: {exc}")

    if errors:
        raise ExcelImportError(
            "Some phone numbers in the Excel file are invalid.",
            errors[:20] + (["..."] if len(errors) > 20 else []),
        )

    if not numbers:
        raise ExcelImportError("No valid phone numbers were found in the Excel file.")

    deduped: list[str] = []
    seen: set[str] = set()
    for number in numbers:
        if number not in seen:
            seen.add(number)
            deduped.append(number)

    return deduped
